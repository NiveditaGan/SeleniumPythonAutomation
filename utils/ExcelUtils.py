import openpyxl
from openpyxl.cell import cell

class TestGetData():
    @staticmethod
    def getCellData(testCaseName):
       # for testCaseName in listofData:

        Dict = {}
        book = openpyxl.load_workbook("/Users/niveditagan/PycharmProjects/PythonFramework/testData/TestData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            val=sheet.cell(row=i, column=1).value
            print(sheet.cell(row=i, column=1))
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(1, sheet.max_column + 1):  # to get through columns if searched testcase name matches
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
        return[Dict]

    @staticmethod
    def putCellData(testCaseName):
        pass